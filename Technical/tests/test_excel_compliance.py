"""Test Excel file compliance with the one-sheet standard."""
import pytest
import openpyxl
from pathlib import Path


def collect_excel_files(directory: Path):
    """Collect all Excel files in directory recursively."""
    if not directory.exists():
        return []
    return list(directory.rglob("*.xlsx"))


class TestExcelCompliance:
    """Verify all Excel files have exactly one sheet."""

    def test_output_data_single_sheet(self, output_data_dir):
        """All Excel files in Output/Data/ must have exactly one sheet."""
        files = collect_excel_files(output_data_dir)
        assert len(files) > 0, "No Excel files found in Output/Data/"

        violations = []
        for f in files:
            try:
                wb = openpyxl.load_workbook(f, read_only=True)
                if len(wb.sheetnames) != 1:
                    violations.append(f"{f.name}: {len(wb.sheetnames)} sheets")
                wb.close()
            except Exception as e:
                violations.append(f"{f.name}: ERROR - {e}")

        assert not violations, (
            f"Excel one-sheet violations:\n" + "\n".join(f"  - {v}" for v in violations)
        )

    def test_country_excel_single_sheet_sample(self, countries_dir):
        """Sample of country Excel files must have exactly one sheet."""
        all_files = collect_excel_files(countries_dir)
        assert len(all_files) > 0, "No country Excel files found"

        # Sample every 10th file for speed
        sample = all_files[::10]

        violations = []
        for f in sample:
            try:
                wb = openpyxl.load_workbook(f, read_only=True)
                if len(wb.sheetnames) != 1:
                    violations.append(f"{f.relative_to(countries_dir)}: {len(wb.sheetnames)} sheets")
                wb.close()
            except Exception as e:
                violations.append(f"{f.name}: ERROR - {e}")

        assert not violations, (
            f"Country Excel violations ({len(sample)} sampled):\n"
            + "\n".join(f"  - {v}" for v in violations)
        )

    def test_output_excel_not_empty(self, output_data_dir):
        """Excel files in Output/Data/ should not be empty."""
        import pandas as pd

        files = list(output_data_dir.glob("*.xlsx"))
        empty_files = []

        for f in files:
            try:
                df = pd.read_excel(f)
                if len(df) == 0:
                    empty_files.append(f.name)
            except Exception:
                pass  # Read errors caught elsewhere

        assert not empty_files, f"Empty Excel files: {empty_files}"
